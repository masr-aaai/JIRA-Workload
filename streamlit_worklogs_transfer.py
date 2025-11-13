import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# =========================
# Column Configuration
# =========================

# Column names inside the worklog files
RESOURCE_COL = "Resource No."
DATE_COL = "Date"
START_COL = "Start Time"
END_COL = "End Time"
TEXT_COL = "Text/Description"
HOURTYPE_COL = "Hour Type"

# Columns inside Employees.xlsx
EMP_RESOURCE_COL = "Resource No."
EMP_PERSONAL_COL = "Pers.Nr."  # adjust if your file uses different naming

# Constants for eco2ve_TimeSheet
CONST_D = "P.0785215.1.02 "
CONST_E = "AN03"

# First row to fill in the eco2ve_TimeSheet (1-based)
START_ROW = 3


def read_employees(employees_path: Path) -> pd.DataFrame:
    df = pd.read_excel(employees_path, dtype=str)

    missing = []
    for col in [EMP_RESOURCE_COL, EMP_PERSONAL_COL]:
        if col not in df.columns:
            missing.append(col)

    if missing:
        raise ValueError(
            f"The following columns are missing in Employees.xlsx: {', '.join(missing)}"
        )

    return df[[EMP_RESOURCE_COL, EMP_PERSONAL_COL]]


def read_all_worklogs(month_folder: Path) -> pd.DataFrame:
    """Read all .xlsx files inside the monthly folder except templates."""
    xlsx_files = [
        f for f in month_folder.glob("*.xlsx")
        if f.name.lower() not in ["employees.xlsx", "eco2ve_timesheet.xlsx"] and "eco2vetimesheet" not in f.name.lower()
    ]

    if not xlsx_files:
        raise FileNotFoundError("No worklog Excel files were found in this month folder.")

    dfs = []
    for f in xlsx_files:
        df = pd.read_excel(f)

        # Check required columns
        required = [RESOURCE_COL, DATE_COL, START_COL, END_COL, TEXT_COL, HOURTYPE_COL]
        for col in required:
            if col not in df.columns:
                raise ValueError(
                    f"File {f.name} is missing the required column '{col}'. "
                    "Please check this worklog file."
                )

        # Keep only billable hours
        before_count = len(df)
        df = df[df[HOURTYPE_COL].astype(str).str.lower() == "billable"]

        if len(df) < before_count:
            st.warning(f"{f.name}: {before_count - len(df)} entries ignored (not billable)")

        dfs.append(df[[RESOURCE_COL, DATE_COL, START_COL, END_COL, TEXT_COL]])

    combined = pd.concat(dfs, ignore_index=True)
    return combined


def build_timesheet_from_template(
    template_path: Path,
    employees_df: pd.DataFrame,
    worklogs_df: pd.DataFrame,
    output_path: Path,
):
    # Mapping Resource No. → Personal Number
    mapping = dict(
        zip(
            employees_df[EMP_RESOURCE_COL].astype(str),
            employees_df[EMP_PERSONAL_COL].astype(str),
        )
    )

    # Add personal number to worklogs
    worklogs_df = worklogs_df.copy()
    worklogs_df[RESOURCE_COL] = worklogs_df[RESOURCE_COL].astype(str)
    worklogs_df["PersonalNumber"] = worklogs_df[RESOURCE_COL].map(mapping)

    # Warn about missing matches
    missing_mask = worklogs_df["PersonalNumber"].isna()
    missing_resources = sorted(worklogs_df.loc[missing_mask, RESOURCE_COL].unique())
    if missing_resources:
        st.warning(
            "For some Resource No. entries no Personal Number was found in Employees.xlsx:\n"
            + ", ".join(missing_resources)
        )

    # Load template
    wb = load_workbook(template_path)
    ws = wb.active  # use first sheet unless specified otherwise

    current_row = START_ROW

    # Convert Date column
    try:
        worklogs_df[DATE_COL] = pd.to_datetime(worklogs_df[DATE_COL]).dt.date
    except Exception:
        pass

    worklogs_df = worklogs_df.sort_values(by=[DATE_COL, RESOURCE_COL]).reset_index(drop=True)

    for _, row in worklogs_df.iterrows():
        ws[f"A{current_row}"] = row.get("PersonalNumber")  # Column A
        ws[f"B{current_row}"] = row.get(DATE_COL)          # Column B
        ws[f"D{current_row}"] = CONST_D                    # Column D
        ws[f"E{current_row}"] = CONST_E                    # Column E
        ws[f"H{current_row}"] = row.get(START_COL)         # Column H
        ws[f"I{current_row}"] = row.get(END_COL)           # Column I
        ws[f"N{current_row}"] = row.get(TEXT_COL)          # Column N

        current_row += 1

    wb.save(output_path)


def main():
    st.title("eco2ve TimeSheet Generator")
    st.write(
        """
Select a base folder that contains:
- `Employees.xlsx`
- `eco2ve_TimeSheet.xlsx` (template)
- one or more monthly folders (e.g. `2025-10`) containing the worklog .xlsx files.
"""
    )

    base_dir_str = st.text_input(
        "Base folder path (e.g. `C:/Users/you/Documents/eco2ve`):",
        value=""
    )

    if st.button("Generate TimeSheet"):
        if not base_dir_str:
            st.error("Please enter a base folder.")
            return

        base_dir = Path(base_dir_str).expanduser()

        if not base_dir.exists() or not base_dir.is_dir():
            st.error("The entered path does not exist or is not a folder.")
            return

        employees_path = base_dir / "Employees.xlsx"
        template_path = base_dir / "eco2ve_TimeSheet.xlsx"

        if not employees_path.exists():
            st.error(f"`Employees.xlsx` was not found at: {employees_path}")
            return

        if not template_path.exists():
            st.error(f"`eco2ve_TimeSheet.xlsx` was not found at: {template_path}")
            return

        # Detect current month automatically
        current_month_str = datetime.today().strftime("%Y-%m")
        month_folder = base_dir / current_month_str

        if not month_folder.exists() or not month_folder.is_dir():
            st.error(
                f"The monthly folder `{current_month_str}` does not exist here: {month_folder}"
            )
            return

        st.info(f"Using monthly folder: `{month_folder}`")

        try:
            employees_df = read_employees(employees_path)
            worklogs_df = read_all_worklogs(month_folder)

            output_filename = f"{current_month_str}-eco2veTimeSheet.xlsx"
            output_path = month_folder / output_filename

            build_timesheet_from_template(
                template_path=template_path,
                employees_df=employees_df,
                worklogs_df=worklogs_df,
                output_path=output_path,
            )

            st.success(f"TimeSheet has been generated: {output_path}")

            # Provide download button
            with open(output_path, "rb") as f:
                st.download_button(
                    label="Download generated file",
                    data=f,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        except Exception as e:
            st.error(f"Processing error: {e}")


if __name__ == "__main__":
    main()
