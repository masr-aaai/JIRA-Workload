# streamlit_jira_worklogs_public.py
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Public Streamlit UI for Jira Worklogs (REST API v3, enhanced /rest/api/3/search/jql)

Features
- Login form (Base URL, Email, API Token, Time zone) — nothing persisted; kept only in session RAM
- Tab A: Month -> XLSX with one sheet per employee (Weekly/Monthly grain toggle)
- Tab B: Employee + Time range -> XLSX for that employee (sheets per Month or per ISO Week)
- Security: no secret logging, masked error details, no caching of credentials, in-memory downloads, Logout

Dependencies:
  pip install streamlit requests pandas openpyxl
"""
import io
from datetime import datetime, timezone, timedelta, date
from typing import Dict, Any, List, Tuple, Optional

import requests
import pandas as pd
import streamlit as st
from requests.auth import HTTPBasicAuth

# zoneinfo (stdlib on Py ≥3.9; fallback to backports for <3.9)
try:
    from zoneinfo import ZoneInfo
except Exception:  # pragma: no cover
    from backports.zoneinfo import ZoneInfo  # pip install backports.zoneinfo

# ------------------------ Security / UI setup ------------------------
st.set_page_config(page_title="Jira Worklogs → Excel", page_icon="📊", layout="centered")
# Hide detailed tracebacks in the browser
st.set_option("client.showErrorDetails", False)

# Masking helper to avoid printing secrets anywhere
def redact(text: str, secrets: List[str]) -> str:
    try:
        txt = str(text)
    except Exception:
        txt = "<unprintable>"
    for s in secrets:
        if s:
            txt = txt.replace(s, "******")
    return txt

# ------------------------ Date & utils ------------------------
WEEKDAY_EN = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

def month_bounds(ym: str) -> Tuple[datetime, datetime]:
    # ym: "YYYY-MM"
    start = datetime.strptime(ym + "-01", "%Y-%m-%d").replace(tzinfo=timezone.utc)
    nextm = start.replace(year=start.year + 1, month=1) if start.month == 12 else start.replace(month=start.month + 1)
    end = nextm - timedelta(seconds=1)
    return start, end

def month_iter(from_ym: str, to_ym: str) -> List[str]:
    sf = datetime.strptime(from_ym + "-01", "%Y-%m-%d")
    stp = datetime.strptime(to_ym + "-01", "%Y-%m-%d")
    if sf > stp:
        sf, stp = stp, sf
    months = []
    y, m = sf.year, sf.month
    while (y < stp.year) or (y == stp.year and m <= stp.month):
        months.append(f"{y:04d}-{m:02d}")
        y, m = (y + 1, 1) if m == 12 else (y, m + 1)
    return months

def parse_started(ts: str) -> datetime:
    for f in ("%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S%z"):
        try:
            return datetime.strptime(ts, f)
        except ValueError:
            pass
    raise ValueError(f"Unknown timestamp format in worklog.started: {ts}")

def safe_sheet_name(s: str) -> str:
    for ch in ['\\','/','*','?','[',']',':']:
        s = s.replace(ch, ' ')
    return (s or "Unknown")[:31]

def week_key_local(dt_local: datetime) -> str:
    # ISO week label like 2025-W05
    iso = dt_local.isocalendar()  # (year, week, weekday)
    return f"{iso[0]:04d}-W{iso[1]:02d}"

# --- Debug controls & helpers (put this once, near the sidebar/auth code) ---

# Toggle in der Sidebar
with st.sidebar:
    st.markdown("### 🛠 Debug")
    st.session_state["DEBUG_ON"] = st.checkbox(
        "Enable debug logs",
        value=st.session_state.get("DEBUG_ON", False)
    )

def _get_debug_flag() -> bool:
    return bool(st.session_state.get("DEBUG_ON", False))

def dlog(msg, *args):
    """Write debug messages to the sidebar when DEBUG is on."""
    if _get_debug_flag():
        try:
            st.sidebar.write("🔎", msg.format(*args) if args else msg)
        except Exception:
            pass

# Falls du 'redact' nicht im File hast, nimm diese Minimalversion:
def redact(text: str, secrets) -> str:
    try:
        s = str(text)
    except Exception:
        return text
    for sec in secrets or []:
        if sec:
            s = s.replace(sec, "******")
    return s

def _redact_dbg(text: str) -> str:
    creds = st.session_state.get("creds", {})
    return redact(text, [creds.get("email",""), creds.get("token",""), creds.get("base","")])

def build_summary_from_month_rows(month_rows: Dict[str, List[Dict[str, Any]]]) -> pd.DataFrame:
    """
    Summary: Stunden pro Fix Version & Scope (Planned/Unplanned) je Monat.
    Erwartet in Monats-Rows: 'Fix Version', 'Scope', 'h logged'
    """
    months_sorted = sorted(month_rows.keys())
    agg: Dict[Tuple[str, str], Dict[str, float]] = {}

    def norm_scope(s: str) -> Optional[str]:
        if not s:
            return None
        s2 = str(s).strip()
        if s2 == "Planned":
            return "Planned"
        if s2 == "Unplanned":
            return "Unplanned"
        return None

    for m in months_sorted:
        rows = month_rows.get(m, [])
        if not rows:
            continue
        dfm = pd.DataFrame(rows)
        if dfm.empty or any(c not in dfm.columns for c in ["h logged","Scope","Fix Version"]):
            continue

        hours = pd.to_numeric(dfm["h logged"], errors="coerce").fillna(0.0).astype(float)
        version = dfm["Fix Version"].fillna("").astype(str)
        scope   = dfm["Scope"].fillna("").astype(str)

        scope_norm = scope.apply(norm_scope)
        mask = scope_norm.notna() & hours.notna()

        temp = pd.DataFrame({
            "version": version[mask],
            "scope": scope_norm[mask],
            "hours": hours[mask],
        })
        grouped = temp.groupby(["version","scope"], dropna=False)["hours"].sum()
        for (ver, sc), hsum in grouped.items():
            agg.setdefault((ver, sc), {})[m] = float(hsum)

    if not agg:
        return pd.DataFrame(columns=["Version - Scope"] + months_sorted)

    keys_sorted = sorted(agg.keys(), key=lambda x: (x[0], x[1]))
    rows_out = []
    for (ver, sc) in keys_sorted:
        row = {"Version - Scope": f"{ver} - {sc}"}
        for m in months_sorted:
            row[m] = round(agg.get((ver, sc), {}).get(m, 0.0), 2)
        rows_out.append(row)

    return pd.DataFrame(rows_out, columns=["Version - Scope"] + months_sorted)

# ------------------------ Jira HTTP ------------------------
def jira_req(base: str, email: str, token: str, method: str, path: str, **kwargs):
    url = f"{base.rstrip('/')}{path}"
    try:
        r = requests.request(
            method, url,
            auth=HTTPBasicAuth(email, token),
            headers={"Accept": "application/json", "Content-Type": "application/json"},
            timeout=90,
            **kwargs
        )
        r.raise_for_status()
        return r.json() if r.text.strip() else {}
    except requests.HTTPError as e:
        detail = r.text if 'r' in locals() else str(e)
        # Mask secrets + base URL
        raise RuntimeError(redact(f"HTTP {getattr(r, 'status_code', '?')} at {url}\n{detail}", [email, token, base])) from e

def search_issues_jql(base, email, token, jql: str,
                      max_results: int = 100,
                      fields: Optional[List[str]] = None,
                      expand: Optional[str] = None) -> List[Dict[str, Any]]:
    """Enhanced search: POST /rest/api/3/search/jql with nextPageToken pagination"""
    def _post(next_token=None):
        payload = {"jql": jql, "maxResults": max_results}
        if fields: payload["fields"] = fields
        if expand: payload["expand"] = expand
        if next_token: payload["nextPageToken"] = next_token
        return jira_req(base, email, token, "POST", "/rest/api/3/search/jql", json=payload)

    issues, next_tok = [], None
    while True:
        page = _post(next_tok)
        issues.extend(page.get("issues", []))
        if page.get("isLast", True): break
        next_tok = page.get("nextPageToken")
        if not next_tok: break
    return issues

def collect_issue_worklogs(base, email, token, issue_key: str) -> List[Dict[str, Any]]:
    """GET /rest/api/3/issue/{key}/worklog (paginated)"""
    all_logs, start_at = [], 0
    while True:
        data = jira_req(
            base, email, token, "GET",
            f"/rest/api/3/issue/{issue_key}/worklog",
            params={"startAt": start_at, "maxResults": 100}
        )
        logs = data.get("worklogs", []) or []
        all_logs.extend(logs)
        total = data.get("total", len(all_logs))
        start_at += len(logs)
        if start_at >= total or not logs:
            break
    return all_logs

def extract_worklog_datetimes(wl: Dict[str, Any], tzinfo_local):
    """
    Liefert:
      - started_utc / started_local: Startzeit des Worklogs (falls vorhanden)
      - created_utc / created_local: Zeitpunkt, wann die Buchung angelegt wurde (falls vorhanden)
      - effective_utc / effective_local: 'maßgebliches' Datum zum Gruppieren/Bucketn
        (Prefer started*, sonst created*)
    Alle Werte können None sein, wenn das jeweilige Feld fehlt oder unparsbar.
    """
    started_raw = wl.get("started")
    created_raw = wl.get("created") or wl.get("updated")  # created bevorzugen, sonst zumindest updated

    started_utc = started_local = created_utc = created_local = None
    if started_raw:
        try:
            started_utc = parse_started(started_raw)
            started_local = started_utc.astimezone(tzinfo_local)
        except Exception:
            started_utc = started_local = None
    if created_raw:
        try:
            created_utc = parse_started(created_raw)
            created_local = created_utc.astimezone(tzinfo_local)
        except Exception:
            created_utc = created_local = None

    effective_utc = started_utc or created_utc
    effective_local = started_local or created_local
    return started_utc, started_local, created_utc, created_local, effective_utc, effective_local

# --- Epic resolver with inheritance (team-managed + company-managed) ---
# Requires: jira_req(...) already defined

from typing import Optional, Tuple, Dict, Any

def resolve_epic_for_issue(
    base: str,
    email: str,
    token: str,
    issue: Dict[str, Any],
    epic_cache: Dict[str, Optional[str]],
    issue_epic_cache: Optional[Dict[str, Tuple[Optional[str], Optional[str]]]] = None,
    ) -> Tuple[Optional[str], Optional[str]]:
    """
    Returns (epic_key, epic_summary) for an issue with proper inheritance:
      1) Team-managed: fields.epic {key|id, name|summary}
      2) Company-managed: Epic Link customfield (commonly customfield_10014)
      3) Parent fallback:
         - if parent is Epic -> use parent
         - else resolve parent's epic (one hop)
    Caches:
      - epic_cache: epic_key -> summary
      - issue_epic_cache: issue_key -> (epic_key, epic_summary)
    """
    if issue_epic_cache is None:
        issue_epic_cache = {}

    ikey = issue.get("key")
    if ikey in issue_epic_cache:
        return issue_epic_cache[ikey]

    f = issue.get("fields") or {}

    def ensure_epic_summary(epic_key: str) -> Optional[str]:
        if not epic_key:
            return None
        if epic_key in epic_cache:
            return epic_cache[epic_key]
        try:
            data = jira_req(base, email, token, "GET", f"/rest/api/3/issue/{epic_key}", params={"fields": "summary"})
            s = (data.get("fields") or {}).get("summary")
        except Exception:
            s = None
        epic_cache[epic_key] = s
        return s

    # 1) Team-managed direct epic object
    epic_obj = f.get("epic")
    if isinstance(epic_obj, dict):
        ekey = epic_obj.get("key") or epic_obj.get("id")
        esum = epic_obj.get("name") or epic_obj.get("summary") or ensure_epic_summary(ekey)
        issue_epic_cache[ikey] = (ekey, esum)
        return ekey, esum

    # 2) Company-managed Epic Link custom field (default id often customfield_10014)
    ekey = f.get("customfield_10014")
    if ekey:
        esum = ensure_epic_summary(ekey)
        issue_epic_cache[ikey] = (ekey, esum)
        return ekey, esum

    # 3) Inherit from parent (subtasks etc.)
    parent = f.get("parent")
    if isinstance(parent, dict) and parent.get("key"):
        pkey = parent.get("key")
        pfields = parent.get("fields") or {}
        ptype = (pfields.get("issuetype") or {}).get("name")

        # Parent is Epic
        if ptype == "Epic":
            esum = pfields.get("summary")
            issue_epic_cache[ikey] = (pkey, esum)
            return pkey, esum

        # Parent might itself have an epic (use embedded data first)
        pepic = pfields.get("epic")
        if isinstance(pepic, dict):
            e2 = pepic.get("key") or pepic.get("id")
            s2 = pepic.get("name") or pepic.get("summary") or ensure_epic_summary(e2)
            issue_epic_cache[ikey] = (e2, s2)
            return e2, s2

        e2 = pfields.get("customfield_10014")
        if e2:
            s2 = ensure_epic_summary(e2)
            issue_epic_cache[ikey] = (e2, s2)
            return e2, s2

        # If still unknown, fetch parent once and inspect
        if pkey in issue_epic_cache:
            e2, s2 = issue_epic_cache[pkey]
            issue_epic_cache[ikey] = (e2, s2)
            return e2, s2

        try:
            pdata = jira_req(base, email, token, "GET", f"/rest/api/3/issue/{pkey}",
                             params={"fields": "issuetype,summary,epic,customfield_10014"})
            pf = (pdata.get("fields") or {})
            if ((pf.get("issuetype") or {}).get("name")) == "Epic":
                e2, s2 = pkey, pf.get("summary")
            else:
                pep = pf.get("epic")
                if isinstance(pep, dict):
                    e2 = pep.get("key") or pep.get("id")
                    s2 = pep.get("name") or pep.get("summary") or ensure_epic_summary(e2)
                else:
                    e2 = pf.get("customfield_10014")
                    s2 = ensure_epic_summary(e2) if e2 else None
            issue_epic_cache[pkey] = (e2, s2)
            issue_epic_cache[ikey] = (e2, s2)
            return e2, s2
        except Exception:
            pass

    # Nothing found
    issue_epic_cache[ikey] = (None, None)
    return None, None

def find_custom_field_id(base: str, email: str, token: str, name_candidates=None) -> Optional[str]:
    """
    Holt /rest/api/3/field und sucht das Scope-Feld über Namen/Klauseln.
    Ändert NICHT die JQL-Query; wir erweitern nur die 'fields' Auswahl.
    """
    if name_candidates is None:
        name_candidates = ["Scope", "scope[dropdown]"]

    try:
        fields = jira_req(base, email, token, "GET", "/rest/api/3/field")
    except Exception as e:
        dlog("[scope] Could not list fields: {}", _redact_dbg(e))
        return None

    want = {s.strip().lower() for s in name_candidates if s and str(s).strip()}
    best_id = None
    for fld in fields:
        fid = fld.get("id")
        if not fid:
            continue
        fname = (fld.get("name") or "").strip().lower()
        clauses = [c.strip().lower() for c in (fld.get("clauseNames") or []) if isinstance(c, str)]
        if fname in want or any(c in want for c in clauses):
            return fid
        if not best_id and "scope" in fname:
            best_id = fid
    return best_id


#------------------------- Excel Formatting ------------------------

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def _find_col_idx_by_header(ws, header_name: str) -> int:
    for cell in ws[1]:
        if (cell.value or "").strip().lower() == header_name.strip().lower():
            return cell.column  # 1-based
    return -1

def apply_conditional_formatting_tab3(ws):
    """
    Einzige Regel:
      - GANZE ZEILE hellgrün, wenn Issue Type == "Epic"
    """
    max_row = ws.max_row
    max_col = ws.max_column
    if max_col < 1 or max_row < 2:
        # Freeze+Filter trotzdem setzen
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"
        ws.freeze_panes = "A2"
        return

    col_type = _find_col_idx_by_header(ws, "Issue Type")
    if col_type > 0:
        green = PatternFill(start_color="FFE6FFCC", end_color="FFE6FFCC", fill_type="solid")
        type_letter = get_column_letter(col_type)
        ws.conditional_formatting.add(
            f"A2:{get_column_letter(max_col)}{max_row}",
            FormulaRule(formula=[f'${type_letter}2="Epic"'], fill=green)
        )

    # Header fixieren + AutoFilter
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"
    ws.freeze_panes = "A2"


# ------------------------ Exports ------------------------
def export_month_all_staff(base, email, token, ym: str, jql_scope: Optional[str], local_tz: str, grain: str) -> bytes:
    """
    grain: 'Monthly' or 'Weekly'
    - Monthly: one sheet per employee (+ Overview)
    - Weekly:   one sheet per employee (+ Weekly Overview)
    Bucket-Key basiert auf dem Starting date (falls vorhanden), sonst created.
    """
    tzinfo_local = ZoneInfo(local_tz)
    start, end = month_bounds(ym)
    date_from = start.astimezone(timezone.utc).strftime("%Y-%m-%d")
    date_to = end.astimezone(timezone.utc).strftime("%Y-%m-%d")

    jql = f'worklogDate >= "{date_from}" AND worklogDate <= "{date_to}"'
    if jql_scope:
        jql = f"({jql_scope}) AND ({jql})"

    issues = search_issues_jql(base, email, token, jql=jql, fields=["summary"])
    by_person: Dict[str, List[Dict[str, Any]]] = {}

    for it in issues:
        key = it.get("key")
        summary = (it.get("fields") or {}).get("summary")
        for wl in collect_issue_worklogs(base, email, token, key):
            # Alle relevanten Zeiten ziehen
            (started_utc, started_local,
             created_utc, created_local,
             effective_utc, effective_local) = extract_worklog_datetimes(wl, tzinfo_local)

            # Filtern: auf Monatsfenster anhand des Starting date (oder created, falls started fehlt)
            if not effective_utc:
                continue
            if not (start <= effective_utc <= end):
                continue

            secs = wl.get("timeSpentSeconds") or 0
            comment = wl.get("comment", "")
            if isinstance(comment, dict):
                comment = comment.get("content") or ""

            row = {
                # Sichtbarkeit der Start- und Buchungszeitpunkte
                "Started (local)": started_local.strftime("%Y-%m-%d %H:%M:%S") if started_local else "",
                "Started (UTC)": started_utc.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S") if started_utc else "",
                "Logged (local)": created_local.strftime("%Y-%m-%d %H:%M:%S") if created_local else "",
                "Logged (UTC)": created_utc.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S") if created_utc else "",
                # Gruppierungs-/Darstellungszeitpunkt
                "Local datetime": effective_local.strftime("%Y-%m-%d %H:%M:%S"),
                "Weekday": WEEKDAY_EN[effective_local.weekday()],
                "ISO Week": week_key_local(effective_local),
                "UTC datetime": effective_utc.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
                # Kontext
                "Issue": key,
                "Summary": summary,
                "Seconds": secs,
                "Hours": round(secs / 3600.0, 2),
                "Comment": comment if isinstance(comment, str) else str(comment),
            }
            author = (wl.get("author") or {})
            person = author.get("displayName") or author.get("accountId") or "Unknown"
            by_person.setdefault(person, []).append(row)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        overview = []
        weekly_overview_rows = []
        for person, rows in by_person.items():
            df = pd.DataFrame(rows).sort_values(["Local datetime", "Issue"])
            df.to_excel(writer, index=False, sheet_name=safe_sheet_name(person))
            overview.append({
                "Employee": person,
                "Entries": len(df),
                "Total Hours": round(float(df["Hours"].sum()), 2) if not df.empty else 0.0
            })
            if grain == "Weekly" and not df.empty:
                wk = df.groupby("ISO Week", dropna=False)["Hours"].sum().reset_index()
                for _, r in wk.iterrows():
                    weekly_overview_rows.append({
                        "Employee": person,
                        "ISO Week": r["ISO Week"],
                        "Total Hours": round(float(r["Hours"]), 2)
                    })

        if overview:
            pd.DataFrame(overview).sort_values("Employee").to_excel(writer, index=False, sheet_name="Overview")
        if grain == "Weekly" and weekly_overview_rows:
            pd.DataFrame(weekly_overview_rows).sort_values(["Employee", "ISO Week"]).to_excel(writer, index=False, sheet_name="Weekly Overview")
        # --- Summary-Sheet bauen & schreiben ---
        summary_df = build_summary_from_month_rows(summary_rows_by_month)

        # existierendes Summary-Blatt ggf. entfernen (falls vorher leer angelegt)
        try:
            if "Summary" in writer.book.sheetnames:
                # openpyxl: Sheet löschen, bevor wir neu schreiben
                ws_old = writer.book["Summary"]
                writer.book.remove(ws_old)
        except Exception:
            pass

        # schreiben
        summary_df.to_excel(writer, index=False, sheet_name="Summary")

        # kleines UX-Setup
        try:
            ws_sum = writer.book["Summary"]
        except Exception:
            ws_sum = writer.sheets.get("Summary")
        if ws_sum is not None:
            from openpyxl.utils import get_column_letter
            max_col = ws_sum.max_column
            ws_sum.freeze_panes = "A2"
            ws_sum.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"

    output.seek(0)
    return output.read()

def list_employees_in_range(base, email, token, from_ym: str, to_ym: str, jql_scope: Optional[str]) -> Dict[str, Dict[str, str]]:
    start_all, end_all = month_bounds(from_ym)[0], month_bounds(to_ym)[1]
    date_from = start_all.astimezone(timezone.utc).strftime("%Y-%m-%d")
    date_to = end_all.astimezone(timezone.utc).strftime("%Y-%m-%d")
    jql = f'worklogDate >= "{date_from}" AND worklogDate <= "{date_to}"'
    if jql_scope: jql = f"({jql_scope}) AND ({jql})"

    emp: Dict[str, Dict[str, str]] = {}
    issues = search_issues_jql(base, email, token, jql=jql, fields=["summary"])
    for it in issues:
        key = it.get("key")
        for wl in collect_issue_worklogs(base, email, token, key):
            try:
                dt_utc = parse_started(wl.get("started"))
            except Exception:
                continue
            if not (start_all <= dt_utc <= end_all):
                continue
            author = (wl.get("author") or {})
            acc = author.get("accountId")
            name = author.get("displayName") or acc or "Unknown"
            if acc and acc not in emp:
                emp[acc] = {"accountId": acc, "displayName": name}
    return emp

def export_employee_timeline(base, email, token, account_id: str,
                             from_ym: str, to_ym: str,
                             jql_scope: Optional[str], local_tz: str,
                             grain: str) -> bytes:
    """
    For a single employee:
      - Monthly grain:  sheet per YYYY-MM
      - Weekly grain:   sheet per ISO week (YYYY-Www)
    Bucket-Key basiert auf dem Starting date (falls vorhanden), sonst created.
    """
    tzinfo_local = ZoneInfo(local_tz)
    start_all, end_all = month_bounds(from_ym)[0], month_bounds(to_ym)[1]
    date_from = start_all.astimezone(timezone.utc).strftime("%Y-%m-%d")
    date_to = end_all.astimezone(timezone.utc).strftime("%Y-%m-%d")
    jql = f'worklogDate >= "{date_from}" AND worklogDate <= "{date_to}"'
    if jql_scope:
        jql = f"({jql_scope}) AND ({jql})"

    issues = search_issues_jql(base, email, token, jql=jql, fields=["summary"])

    # Prepare buckets (by Month or ISO Week) – based on effective_local
    if grain == "Monthly":
        buckets: Dict[str, List[Dict[str, Any]]] = {ym: [] for ym in month_iter(from_ym, to_ym)}
        def bucket_key(dt_local: datetime) -> str:
            return f"{dt_local.year:04d}-{dt_local.month:02d}"
    else:
        buckets = {}
        def bucket_key(dt_local: datetime) -> str:
            return week_key_local(dt_local)

    for it in issues:
        key = it.get("key")
        summary = (it.get("fields") or {}).get("summary")
        for wl in collect_issue_worklogs(base, email, token, key):
            (started_utc, started_local,
             created_utc, created_local,
             effective_utc, effective_local) = extract_worklog_datetimes(wl, tzinfo_local)

            # Zeitraumfilter anhand effective_utc (started bevorzugt)
            if not effective_utc or not (start_all <= effective_utc <= end_all):
                continue
            author = (wl.get("author") or {})
            if author.get("accountId") != account_id:
                continue

            secs = wl.get("timeSpentSeconds") or 0
            comment = wl.get("comment", "")
            if isinstance(comment, dict):
                comment = comment.get("content") or ""

            row = {
                "Started (local)": started_local.strftime("%Y-%m-%d %H:%M:%S") if started_local else "",
                "Started (UTC)": started_utc.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S") if started_utc else "",
                "Logged (local)": created_local.strftime("%Y-%m-%d %H:%M:%S") if created_local else "",
                "Logged (UTC)": created_utc.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S") if created_utc else "",
                "Local datetime": effective_local.strftime("%Y-%m-%d %H:%M:%S"),
                "Weekday": WEEKDAY_EN[effective_local.weekday()],
                "ISO Week": week_key_local(effective_local),
                "UTC datetime": effective_utc.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
                "Issue": key,
                "Summary": summary,
                "Seconds": secs,
                "Hours": round(secs / 3600.0, 2),
                "Comment": comment if isinstance(comment, str) else str(comment),
            }
            buckets.setdefault(bucket_key(effective_local), []).append(row)

    # Build workbook in-memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        overview_rows = []
        label = "Month" if grain == "Monthly" else "ISO Week"
        for bkey in sorted(buckets.keys()):
            df = pd.DataFrame(buckets[bkey])
            if not df.empty:
                df = df.sort_values(["Local datetime", "Issue"])
            else:
                df = pd.DataFrame(columns=[
                    "Started (local)","Started (UTC)","Logged (local)","Logged (UTC)",
                    "Local datetime","Weekday","ISO Week","UTC datetime","Issue",
                    "Summary","Seconds","Hours","Comment"
                ])
            df.to_excel(writer, index=False, sheet_name=safe_sheet_name(bkey))
            total_h = round(float(df["Hours"].sum()), 2) if not df.empty else 0.0
            overview_rows.append({label: bkey, "Entries": len(df), "Total Hours": total_h})

        if overview_rows:
            pd.DataFrame(overview_rows).to_excel(writer, index=False, sheet_name="Overview")
    output.seek(0)
    return output.read()

def export_monthly_hierarchy_by_epic(
    base: str,
    email: str,
    token: str,
    from_ym: str,
    to_ym: str,
    jql_scope: Optional[str],
    local_tz: str,
    epic_link_field: Optional[str] = None,  # ungenutzt
    scope_field: Optional[str] = None,      # falls None: auto-detect
    ) -> bytes:
    """
    Tab 3 – Monats-Export als Baum:
      - 'parent' = ROHER Jira-Parent (fields.parent.key)
      - Anzeige als Baum (Epic -> Task -> Subtask); Export-Parent bleibt roh
      - Fehlende Vorfahren werden mit 0h eingefügt, falls nur im Kind geloggt wurde
      - 'Issue Type' immer
      - 'sum h epic' = Summe des Epic-Subtrees
      - 'Scope (Jira)' = Original aus Scope-Customfield
      - 'Scope' = nur dann füllen, wenn 'Scope (Jira)' leer war (Parent erben; falls dann noch leer & Typ != Epic -> "Planned")
      - 'Fix Version' aus fixVersions (CSV)
      - Summary-Sheet wird fertig mitgeschrieben
      - Formatierung: nur grüne Zeilen, wenn Issue Type == 'Epic'
    """
    tzinfo_local = ZoneInfo(local_tz)

    # Zeitraum
    start_all, _ = month_bounds(from_ym)
    _, end_all   = month_bounds(to_ym)
    date_from = start_all.astimezone(timezone.utc).strftime("%Y-%m-%d")
    date_to   = end_all.astimezone(timezone.utc).strftime("%Y-%m-%d")

    # JQL (unverändert) + optionaler Scopefilter
    jql = f'worklogDate >= "{date_from}" AND worklogDate <= "{date_to}"'
    if jql_scope:
        jql = f"({jql_scope}) AND ({jql})"
    dlog("[tab3] JQL: {}", _redact_dbg(jql))

    # Scope-Customfield-ID (auto)
    scope_field_id = scope_field or find_custom_field_id(base, email, token, ["Scope","scope[dropdown]"])
    if scope_field_id:
        dlog("[tab3] detected scope field id: {}", scope_field_id)
    else:
        dlog("[tab3] scope field id NOT found; 'Scope (Jira)' will be empty")

    # Felderliste (JQL selbst bleibt gleich)
    fields = ["summary", "issuetype", "parent", "epic", "customfield_10014", "fixVersions"]
    if scope_field_id:
        fields.append(scope_field_id)

    issues = search_issues_jql(base, email, token, jql=jql, fields=fields)
    dlog("[tab3] issues fetched: {}", len(issues))

    # Caches
    epic_cache: Dict[str, Optional[str]] = {}
    issue_epic_cache: Dict[str, Tuple[Optional[str], Optional[str]]] = {}

    # Metadaten
    issue_meta: Dict[str, Dict[str, Any]] = {}
    raw_parent_map: Dict[str, Optional[str]] = {}
    issuetype_map: Dict[str, str] = {}
    epic_name_map: Dict[str, str] = {}

    # Metadaten sammeln
    for it in issues:
        key = it.get("key")
        if not key:
            continue
        f = it.get("fields") or {}
        summary = f.get("summary") or ""
        issuetype_name = ((f.get("issuetype") or {}).get("name") or "").strip()

        # roher Parent
        parent_key = None
        if isinstance(f.get("parent"), dict):
            parent_key = f["parent"].get("key")

        # Fix Versions CSV
        fix_versions = ""
        fv = f.get("fixVersions") or []
        if isinstance(fv, list):
            names = []
            for v in fv:
                if isinstance(v, dict):
                    nm = v.get("name")
                    if nm:
                        names.append(str(nm))
                else:
                    names.append(str(v))
            fix_versions = ", ".join(names)

        # Epic vererbt/aufgelöst
        e_key, e_sum = resolve_epic_for_issue(
            base, email, token, it,
            epic_cache=epic_cache,
            issue_epic_cache=issue_epic_cache
        )
        if e_key and e_sum:
            epic_name_map[e_key] = e_sum

        # Scope (Jira) direkt aus dem Customfield (wenn vorhanden)
        scope_jira = ""
        if scope_field_id and scope_field_id in f:
            sv = f.get(scope_field_id)
            if isinstance(sv, dict):
                scope_jira = sv.get("value") or sv.get("name") or ""
            elif isinstance(sv, list):
                parts = []
                for item in sv:
                    if isinstance(item, dict):
                        parts.append(item.get("value") or item.get("name") or "")
                    elif item is not None:
                        parts.append(str(item))
                scope_jira = ", ".join([p for p in parts if p])
            elif sv is not None:
                scope_jira = str(sv)

        issue_meta[key] = {
            "key": key,
            "summary": summary,
            "issuetype": issuetype_name,
            "raw_parent": parent_key,
            "epic_key": e_key,
            "epic_summary": e_sum or "",
            "scope_jira": scope_jira or "",
            "fix_versions": fix_versions,
        }
        raw_parent_map[key] = parent_key
        issuetype_map[key] = issuetype_name

        dlog("[tab3] meta issue={} type={} raw_parent={} epic=({}, {}) scope_jira='{}' fixVersions={}",
             key, issuetype_name, parent_key, e_key, e_sum, scope_jira, fix_versions)

    # Stunden pro Monat/Issue
    per_month_issue_hours: Dict[str, Dict[str, float]] = {}
    months = month_iter(from_ym, to_ym)

    for it in issues:
        key = it.get("key")
        if not key:
            continue
        for wl in collect_issue_worklogs(base, email, token, key):
            (started_utc, started_local,
             created_utc, created_local,
             effective_utc, effective_local) = extract_worklog_datetimes(wl, tzinfo_local)
            if not effective_utc or not effective_local:
                continue
            if not (start_all <= effective_utc <= end_all):
                continue
            month_key = f"{effective_local.year:04d}-{effective_local.month:02d}"
            if month_key not in months:
                continue
            secs = wl.get("timeSpentSeconds") or 0
            h = secs / 3600.0
            per_month_issue_hours.setdefault(month_key, {})
            per_month_issue_hours[month_key][key] = per_month_issue_hours[month_key].get(key, 0.0) + h

    # Workbook
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_rows_by_month: Dict[str, List[Dict[str, Any]]] = {}

        for m in months:
            hours = per_month_issue_hours.get(m, {})
            active = set(hours.keys())
            to_include = set(active)

            # Vorfahren (roh) einziehen
            def include_ancestors(issue_key: str):
                cur = issue_key
                visited = set()
                while True:
                    p = raw_parent_map.get(cur)
                    if not p or p in visited:
                        break
                    to_include.add(p)
                    visited.add(p)
                    cur = p

            for k in list(active):
                include_ancestors(k)

            # Baum-Eltern (nur Anzeige); Export-Parent bleibt roh
            tree_parent: Dict[str, Optional[str]] = {k: raw_parent_map.get(k) for k in list(to_include)}

            # Top → Epic hängen (nur Anzeige), Epic-Knoten ggf. anlegen
            for k in list(active):
                cur = k
                seen = set()
                while raw_parent_map.get(cur) and cur not in seen:
                    seen.add(cur)
                    cur = raw_parent_map[cur]
                top = cur
                ekey = issue_meta.get(k, {}).get("epic_key")
                if ekey:
                    if ekey not in issue_meta:
                        issue_meta[ekey] = {
                            "key": ekey,
                            "summary": epic_name_map.get(ekey, ""),
                            "issuetype": "Epic",
                            "raw_parent": None,
                            "epic_key": None,
                            "epic_summary": "",
                            "scope_jira": "",
                            "fix_versions": "",
                        }
                        raw_parent_map[ekey] = None
                        issuetype_map[ekey] = "Epic"
                    to_include.add(ekey)
                    if issuetype_map.get(top) != "Epic":
                        tree_parent[top] = ekey

            # fehlende Stunden = 0
            for k in to_include:
                hours.setdefault(k, 0.0)

            # Scope ableiten NUR wenn „Scope (Jira)“ leer ist
            scope_jira_for_row: Dict[str, str] = {}
            derived_scope_for_row: Dict[str, str] = {}

            for k in to_include:
                scope_jira_for_row[k] = issue_meta.get(k, {}).get("scope_jira", "") or ""

            for k in list(to_include):
                orig = scope_jira_for_row.get(k, "")
                if orig:
                    derived_scope_for_row[k] = orig  # NICHT überschreiben
                    continue

                # Vererbung über *rohe* Parent-Kette
                cur = raw_parent_map.get(k)
                guard = set([k])
                inherited = ""
                while cur and cur not in guard:
                    inherited = scope_jira_for_row.get(cur, "") or issue_meta.get(cur, {}).get("scope_jira", "") or ""
                    if inherited:
                        break
                    guard.add(cur)
                    cur = raw_parent_map.get(cur)

                if not inherited and issuetype_map.get(k) != "Epic":
                    inherited = "Planned"   # Default nur wenn vorher leer
                derived_scope_for_row[k] = inherited

            # Baumkinder
            children: Dict[Optional[str], List[str]] = {}
            for k in to_include:
                children.setdefault(tree_parent.get(k), []).append(k)
            for p in children:
                children[p].sort()
            roots = sorted(children.get(None, []))

            # Subtree-Summen je Epic
            subtree_total_cache: Dict[str, float] = {}
            def subtree_total(node: str) -> float:
                if node in subtree_total_cache:
                    return subtree_total_cache[node]
                total = float(hours.get(node, 0.0))
                for c in children.get(node, []):
                    total += subtree_total(c)
                subtree_total_cache[node] = total
                return total

            epic_totals: Dict[str, float] = {}
            for n in to_include:
                if issuetype_map.get(n) == "Epic":
                    epic_totals[n] = round(subtree_total(n), 2)

            def epic_ancestor(node: str) -> Optional[str]:
                cur = node
                guard = set()
                while cur and cur not in guard:
                    if issuetype_map.get(cur) == "Epic":
                        return cur
                    guard.add(cur)
                    cur = tree_parent.get(cur)
                return None

            # DFS Ausgabe
            rows: List[Dict[str, Any]] = []
            def walk(node_key: str, depth: int = 0):
                meta = issue_meta.get(node_key, {})
                epic_for_row = epic_ancestor(node_key)
                sum_h_epic = epic_totals.get(epic_for_row) if epic_for_row else None
                rows.append({
                    "Issue": node_key,
                    "Issue Type": issuetype_map.get(node_key, ""),
                    "Summary": ("  " * depth) + (meta.get("summary") or ""),
                    "parent": meta.get("raw_parent") or "",                     # roh (wie Tab 4)
                    "h logged": round(float(hours.get(node_key, 0.0)), 2),
                    "sum h epic": (round(float(sum_h_epic), 2) if sum_h_epic is not None else ""),
                    "Scope (Jira)": (issue_meta.get(node_key, {}).get("scope_jira", "") or ""),
                    "Scope": (derived_scope_for_row.get(node_key, "") or ""),
                    "Fix Version": meta.get("fix_versions", ""),
                })
                for child in children.get(node_key, []):
                    walk(child, depth + 1)

            for r in roots:
                walk(r, 0)

            # Für Summary merken
            summary_rows_by_month[m] = list(rows)

            # Sheet schreiben + nur Epic-Grün formatieren
            df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
                "Issue","Issue Type","Summary","parent","h logged","sum h epic","Scope (Jira)","Scope","Fix Version"
            ])
            sheet_name = safe_sheet_name(m)
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            try:
                ws = writer.book[sheet_name]
            except Exception:
                ws = writer.sheets.get(sheet_name)
            if ws is not None:
                apply_conditional_formatting_tab3(ws)

        # Summary-Blatt erzeugen & schreiben
        summary_df = build_summary_from_month_rows(summary_rows_by_month)

        # existierendes Summary ggf. entfernen
        try:
            if "Summary" in writer.book.sheetnames:
                ws_old = writer.book["Summary"]
                writer.book.remove(ws_old)
        except Exception:
            pass

        summary_df.to_excel(writer, index=False, sheet_name="Summary")

        try:
            ws_sum = writer.book["Summary"]
        except Exception:
            ws_sum = writer.sheets.get("Summary")
        if ws_sum is not None:
            from openpyxl.utils import get_column_letter
            max_col = ws_sum.max_column
            ws_sum.freeze_panes = "A2"
            ws_sum.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"

    output.seek(0)
    return output.read()

# ------------------------ Login / Session ------------------------
def logged_in() -> bool:
    return "creds" in st.session_state and all(
        st.session_state["creds"].get(k) for k in ("base", "email", "token", "tz")
    )

with st.sidebar:
    st.header("🔐 Authentication")
    st.text("Enter your Jira credentials. No data is stored server-side; only in session RAM. Create an API token in https://id.atlassian.com/manage-profile/security/api-tokens.")
    if not logged_in():
        with st.form("login"):
            base = st.text_input("Jira Base URL", placeholder="https://your-domain.atlassian.net")
            email = st.text_input("Email")
            token = st.text_input("API Token", type="password")
            tz = st.text_input("Time zone (IANA)", value="Europe/Vienna", help="e.g., Europe/Vienna, UTC, America/New_York")
            submit = st.form_submit_button("Sign in")
        if submit:
            # Store creds only in session (RAM), never cache, never log
            st.session_state["creds"] = {"base": base.strip(), "email": email.strip(), "token": token, "tz": tz.strip()}
            try:
                st.rerun()  # Streamlit ≥ 1.27
            except AttributeError:
                # Fallback für sehr alte Versionen, in denen st.rerun noch nicht existierte
                st.experimental_rerun()
    else:
        st.success("Signed in")
        if st.button("Logout"):
            st.session_state.clear()
            try:
                st.rerun()  # Streamlit ≥ 1.27
            except AttributeError:
                # Fallback für sehr alte Versionen, in denen st.rerun noch nicht existierte
                st.experimental_rerun()

if not logged_in():
    st.stop()

# For convenience
creds = st.session_state["creds"]
BASE, EMAIL, TOKEN, TZ = creds["base"], creds["email"], creds["token"], creds["tz"]

st.title("📊 Jira Worklogs → Excel")
st.caption("Generate Excel files from Jira worklogs — no credentials stored server-side. Choose Monthly or Weekly granularity.")

tab1, tab2, tab3, tab4 = st.tabs([
    "All employees per month",
    "Employee per time range",
    "Timerange → Hierarchy (Month)",
    "Custom JQL → Excel"
])

# ------------------------ Tab 1: Month → All employees ------------------------
with tab1:
    today = date.today()
    default_month = f"{today.year}-{today.month:02d}"
    month = st.text_input("Month (YYYY-MM)", value=default_month, placeholder="e.g., 2025-09", key="t1_month")
    scope = st.text_input("Optional JQL scope (e.g., project = ABC)", value="", key="t1_scope")
    grain = st.selectbox("Granularity", options=["Monthly", "Weekly"], index=0, key="t1_grain",
                         help="Monthly: one sheet per employee; Weekly: plus ISO week columns & a Weekly Overview sheet.")
    if st.button("Generate Excel (all employees)", key="t1_btn"):
        try:
            datetime.strptime(month + "-01", "%Y-%m-%d")
        except ValueError:
            st.error("Please enter month as YYYY-MM, e.g., 2025-09.")
            st.stop()
        try:
            st.info("Fetching issues and worklogs…")
            xlsx_bytes = export_month_all_staff(
                base=BASE, email=EMAIL, token=TOKEN,
                ym=month.strip(), jql_scope=scope.strip() or None,
                local_tz=TZ, grain=grain
            )
            filename = f"worklogs_all_{month}_{grain.lower()}.xlsx"
            st.success("Done. Your file is ready.")
            st.download_button(
                label="📥 Download Excel",
                data=xlsx_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error("Request failed.")
            with st.expander("Error details"):
                st.code(redact(e, [EMAIL, TOKEN, BASE]))

# ------------------------ Tab 2: Employee + Time range ------------------------
with tab2:
    c1, c2 = st.columns(2)
    with c1:
        from_month = st.text_input("Start month (YYYY-MM)", value=default_month, key="t2_from")
    with c2:
        to_month = st.text_input("End month (YYYY-MM)", value=default_month, key="t2_to")
    scope2 = st.text_input("Optional JQL scope (e.g., project = ABC)", value="", key="t2_scope")

    # Load employees present in the selected range
    if st.button("Load employees from range", key="t2_load"):
        try:
            datetime.strptime(from_month + "-01", "%Y-%m-%d")
            datetime.strptime(to_month + "-01", "%Y-%m-%d")
        except ValueError:
            st.error("Please provide months as YYYY-MM, e.g., 2025-01 and 2025-03.")
            st.stop()
        try:
            st.info("Scanning worklogs to list employees in range…")
            emps = list_employees_in_range(
                base=BASE, email=EMAIL, token=TOKEN,
                from_ym=from_month.strip(), to_ym=to_month.strip(),
                jql_scope=scope2.strip() or None
            )
            if not emps:
                st.warning("No employees with worklogs in the selected period.")
            st.session_state["t2_emps"] = emps
        except Exception as e:
            st.error("Failed to list employees.")
            with st.expander("Error details"):
                st.code(redact(e, [EMAIL, TOKEN, BASE]))

    emps = st.session_state.get("t2_emps", {})
    if emps:
        # Sort by displayName
        options = sorted(emps.values(), key=lambda x: (x.get("displayName") or "").lower())
        labels = [f'{e["displayName"]}  ({e["accountId"][:8]}…)' for e in options]
        idx = st.selectbox("Choose employee", options=list(range(len(options))), format_func=lambda i: labels[i], key="t2_select")
        selected_acc = options[idx]["accountId"]
        selected_name = options[idx]["displayName"]

        grain2 = st.selectbox("Granularity", options=["Monthly", "Weekly"], index=0, key="t2_grain",
                              help="Monthly: sheets per month; Weekly: sheets per ISO week.")
        if st.button("Generate Excel for employee", key="t2_btn"):
            try:
                st.info(f"Generating file for {selected_name}…")
                xlsx_bytes = export_employee_timeline(
                    base=BASE, email=EMAIL, token=TOKEN,
                    account_id=selected_acc,
                    from_ym=from_month.strip(), to_ym=to_month.strip(),
                    jql_scope=scope2.strip() or None,
                    local_tz=TZ, grain=grain2
                )
                fn = f"worklogs_{selected_name.replace(' ', '_')}_{from_month}_{to_month}_{grain2.lower()}.xlsx"
                st.success("Done. Your file is ready.")
                st.download_button(
                    label="📥 Download Excel",
                    data=xlsx_bytes,
                    file_name=fn,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                st.error("Request failed.")
                with st.expander("Error details"):
                    st.code(redact(e, [EMAIL, TOKEN, BASE]))

# ------------------------ Tab 3:  Epic Hier ------------------------
with tab3:
    st.subheader("Hierarchische Monatsübersicht (Epic → Workitem → Subtask)")

    c1, c2 = st.columns(2)
    with c1:
        h_from_month = st.text_input("Start month (YYYY-MM)", value=default_month, key="t3_from")
    with c2:
        h_to_month = st.text_input("End month (YYYY-MM)", value=default_month, key="t3_to")

    h_scope = st.text_input(
        "Optional JQL scope (e.g., project = ABC)",
        value="",
        key="t3_scope"
    )

    st.markdown("**Feld-IDs für Hierarchie / Scope**")
    h_epic_link_field = st.text_input(
        "Epic link field ID (optional, e.g. customfield_10014)",
        value="",
        key="t3_epic_link"
    )
    h_scope_field = st.text_input(
        "Scope custom field ID (optional, e.g. customfield_12345)",
        value="",
        key="t3_scope_field"
    )

    st.caption(
        "Hinweis: "
        "`release version` wird aus dem Feld `fixVersions` gelesen. "
        "`Scope` ist ein Custom Field, dessen ID du oben angeben kannst."
    )

    if st.button("Generate hierarchical Excel (by month)", key="t3_btn"):
        # Format prüfen
        try:
            datetime.strptime(h_from_month + "-01", "%Y-%m-%d")
            datetime.strptime(h_to_month + "-01", "%Y-%m-%d")
        except ValueError:
            st.error("Please provide months as YYYY-MM, e.g., 2025-01 and 2025-03.")
            st.stop()

        try:
            st.info("Building hierarchical monthly overview…")
            xlsx_bytes = export_monthly_hierarchy_by_epic(
                base=BASE,
                email=EMAIL,
                token=TOKEN,
                from_ym=h_from_month.strip(),
                to_ym=h_to_month.strip(),
                jql_scope=h_scope.strip() or None,
                local_tz=TZ,
                epic_link_field=h_epic_link_field.strip() or None,
                scope_field=h_scope_field.strip() or None,
            )
            fn = f"worklogs_hierarchical_{h_from_month}_{h_to_month}.xlsx"
            st.success("Done. Your file is ready.")
            st.download_button(
                label="📥 Download Excel (hierarchical)",
                data=xlsx_bytes,
                file_name=fn,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error("Request failed.")
            with st.expander("Error details"):
                st.code(redact(e, [EMAIL, TOKEN, BASE]))


with tab4:
    st.subheader("Export issues by JQL")

    # Example JQL prefilled with your request
    default_jql = 'fixversion = "Version 2.2" AND status NOT IN (abandonded, Abandoned) ORDER BY parent ASC'
    jql_in = st.text_area("JQL", value=default_jql, height=120,
                          help='Paste any valid JQL. Uses enhanced search (/rest/api/3/search/jql).')

    if st.button("Generate Excel (JQL results)", key="t4_generate"):
        if not jql_in.strip():
            st.warning("Please enter a JQL query.")
            st.stop()
        try:
            st.info("Running JQL and building workbook…")
            xlsx_bytes = export_jql_issues_excel(
                base=BASE, email=EMAIL, token=TOKEN,
                jql=jql_in.strip(), local_tz=TZ
            )
            fn = "jql_results.xlsx"
            st.success("Done. Your file is ready.")
            st.download_button(
                label="📥 Download Excel",
                data=xlsx_bytes,
                file_name=fn,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error("Request failed.")
            with st.expander("Error details"):
                st.code(redact(e, [EMAIL, TOKEN, BASE]))
